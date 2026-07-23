"""
Simulation and post-processing
"""

from dataclasses import dataclass
from typing import Callable, List, Optional

import numpy as np
import pandas as pd

from .materials import Material
from .particles import Particle
from .physics import dEdx_mass
from .units import StoppingUnit, convert_from_mass


@dataclass
class SimSettings:
    x_start: float = 0.0        # mm
    x_stop: float = 1000.0      # mm
    E_cutoff: float = 0.001     # MeV
    max_dx: float = 0.5         # mm
    frac_loss: float = 0.01     # max fractional energy loss per step


@dataclass
class TrackResult:
    """
    Bragg curve result.  Arrays are already trimmed to the physical
    range — they end where dE/dx first drops below 1.5 % of the Bragg
    peak after the peak, which marks the boundary of Bethe-Bloch
    validity (nuclear stopping takes over below that point).

    x_mass    : mass thickness (g/cm^2)
    E         : kinetic energy (MeV)
    dEdx_mass : mass stopping power (MeV cm^2 / g)
    rho       : material density (g/cm^3)
    """
    name: str
    color: str
    x_mass: np.ndarray
    E: np.ndarray
    dEdx_mass: np.ndarray
    rho: float

    @property
    def x_mm(self) -> np.ndarray:
        return self.x_mass / self.rho * 10.0

    def dEdx_in(self, unit: StoppingUnit) -> np.ndarray:
        return convert_from_mass(self.dEdx_mass, unit, self.rho)

    def x_in(self, mass_thickness: bool = False) -> np.ndarray:
        return self.x_mass if mass_thickness else self.x_mm

    @property
    def range_mm(self) -> float:
        return float(self.x_mm[-1])

    @property
    def range_mass(self) -> float:
        return float(self.x_mass[-1])

    def peak_dEdx(self, unit: StoppingUnit) -> float:
        return float(np.nanmax(self.dEdx_in(unit)))


# ---- cutoff helper ----

def _trim_after_peak(x_vals: list, E_vals: list, dEdx_vals: list,
                     threshold_frac: float = 0.015):

    arr = np.array(dEdx_vals, dtype=float)
    finite = arr[np.isfinite(arr)]
    if len(finite) == 0:
        return x_vals, E_vals, dEdx_vals

    peak      = float(np.nanmax(finite))
    threshold = threshold_frac * peak
    peak_idx  = int(np.nanargmax(arr))

    post  = arr[peak_idx:]
    below = np.where(np.isfinite(post) & (post < threshold))[0]

    if len(below) == 0:
        return x_vals, E_vals, dEdx_vals

    cutoff = peak_idx + int(below[0])
    return x_vals[:cutoff], E_vals[:cutoff], list(arr[:cutoff])


# ---- integrator ----

def simulate(mat: Material,
             part: Particle,
             settings: SimSettings,
             progress_cb: Optional[Callable[[float], None]] = None) -> TrackResult:

    if part.E0 is None or part.E0 <= 0:
        raise ValueError(f"{part.name}: initial kinetic energy must be > 0 MeV.")

    mm_to_mass = mat.rho / 10.0
    x_stop_m   = settings.x_stop  * mm_to_mass
    max_dx_m   = settings.max_dx  * mm_to_mass
    x_start_m  = settings.x_start * mm_to_mass

    E0 = float(part.E0)
    x, E = x_start_m, E0
    x_vals, E_vals, dEdx_vals = [x], [E], []

    while x < x_stop_m and E > settings.E_cutoff:
        s1 = dEdx_mass(mat, part, E)
        if s1 <= 0:
            break

        dx = min(max_dx_m, settings.frac_loss * E / s1)
        dx = max(dx, 1e-9)

        E_mid = E - 0.5 * dx * s1
        if E_mid <= 0:
            dEdx_vals.append(s1)
            x += dx; E = 0.0
            x_vals.append(x); E_vals.append(E)
            break

        s_mid = dEdx_mass(mat, part, E_mid)
        if s_mid <= 0:
            dEdx_vals.append(s1)
            x += dx; E = 0.0
            x_vals.append(x); E_vals.append(E)
            break

        E_next = E - dx * s_mid
        dEdx_vals.append(s_mid)
        x += dx
        E = max(E_next, 0.0)
        x_vals.append(x); E_vals.append(E)

        if progress_cb is not None and E0 > 0:
            progress_cb(min(1.0, (E0 - E) / E0))

    if len(dEdx_vals) < len(x_vals):
        dEdx_vals.append(np.nan)

    if progress_cb is not None:
        progress_cb(1.0)

    # ---- trim to physical range ----
    x_vals, E_vals, dEdx_vals = _trim_after_peak(x_vals, E_vals, dEdx_vals)

    return TrackResult(
        name=part.name,
        color=part.color,
        x_mass=np.array(x_vals),
        E=np.array(E_vals),
        dEdx_mass=np.array(dEdx_vals),
        rho=mat.rho,
    )


# ---- intersection finder ----
def find_intersections(a: TrackResult, b: TrackResult,
                       unit: StoppingUnit,
                       mass_thickness: bool = False) -> List[tuple]:
    xa, ya = a.x_in(mass_thickness), a.dEdx_in(unit)
    xb, yb = b.x_in(mass_thickness), b.dEdx_in(unit)

    df_a = pd.DataFrame({"x": xa, "ya": ya})
    df_b = pd.DataFrame({"x": xb, "yb": yb})
    df = (pd.merge(df_a, df_b, on="x", how="outer")
            .sort_values("x").reset_index(drop=True))
    df["ya"] = df["ya"].interpolate(limit_area="inside")
    df["yb"] = df["yb"].interpolate(limit_area="inside")

    diff      = (df["ya"] - df["yb"]).dropna()
    cross_idx = np.where(np.diff(np.sign(diff)) != 0)[0]

    intersections = []
    for loc in cross_idx:
        idx = diff.index[loc]
        try:
            nxt = diff.index[loc + 1]
        except IndexError:
            continue
        x_a, x_b = df.loc[idx, "x"],  df.loc[nxt, "x"]
        y1a, y1b = df.loc[idx, "ya"], df.loc[nxt, "ya"]
        y2a, y2b = df.loc[idx, "yb"], df.loc[nxt, "yb"]
        dy1, dy2 = y1b - y1a, y2b - y2a
        if (dy1 - dy2) != 0:
            t  = (y2a - y1a) / (dy1 - dy2)
            ix = x_a + t * (x_b - x_a)
            iy = y1a + t * dy1
            intersections.append((ix, iy))
    return intersections


# ---- export ----

def _build_export_df(results: List[TrackResult], unit: StoppingUnit,
                     mass_thickness: bool) -> pd.DataFrame:
    """Shared table-building logic for CSV and Excel export."""
    x_label = "x_g_per_cm2" if mass_thickness else "x_mm"
    df = None
    for r in results:
        d = pd.DataFrame({
            x_label:           r.x_in(mass_thickness),
            f"E_{r.name}_MeV": r.E,
            f"dEdx_{r.name}":  r.dEdx_in(unit),
        })
        df = d if df is None else pd.merge(df, d, on=x_label, how="outer")
    df = df.sort_values(x_label).reset_index(drop=True)
    for r in results:
        col = f"dEdx_{r.name}"
        if col in df.columns:
            df[col] = df[col].interpolate(limit_area="inside")
    return df


def export_csv(results: List[TrackResult], path: str,
               unit: StoppingUnit, mass_thickness: bool = False) -> None:
    _build_export_df(results, unit, mass_thickness).to_csv(path, index=False)


def export_xlsx(results: List[TrackResult], path: str,
                unit: StoppingUnit, mass_thickness: bool = False) -> None:
    """Same table as export_csv, written as .xlsx via pandas/openpyxl."""
    _build_export_df(results, unit, mass_thickness).to_excel(path, index=False)

def export_xlsx(results: List[TrackResult], path: str,
                unit: StoppingUnit, mass_thickness: bool = False) -> None:
    """
    Sheet "Bragg curves"
        One shared x-column, then alternating energy / dE/dx columns per
        particle — same structure as the CSV export so the two are
        interchangeable.

    Sheet "Metadata"
        Range, peak dE/dx, and unit information for each particle.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows

    # ---- build the merged data frame ----
    x_label = "x_g_per_cm2" if mass_thickness else "x_mm"
    df = None
    for r in results:
        d = pd.DataFrame({
            x_label: r.x_in(mass_thickness),
            f"E_{r.name}_MeV": r.E,
            f"dEdx_{r.name}_{unit.label}": r.dEdx_in(unit),
        })
        df = d if df is None else pd.merge(df, d, on=x_label, how="outer")

    df = df.sort_values(x_label).reset_index(drop=True)
    for r in results:
        col = f"dEdx_{r.name}_{unit.label}"
        if col in df.columns:
            df[col] = df[col].interpolate(limit_area="inside")

    # ---- build metadata frame ----
    meta_rows = []
    for r in results:
        meta_rows.append({
            "Particle": r.name,
            "Range (mm)": round(r.range_mm, 5),
            "Range (g/cm²)": round(r.range_mass, 6),
            f"Peak dE/dx ({unit.label})": round(r.peak_dEdx(unit), 5),
            "dE/dx unit": unit.label,
            "x-axis": "Mass thickness (g/cm²)" if mass_thickness else "Distance (mm)",
        })
    df_meta = pd.DataFrame(meta_rows)

    # ---- write workbook ----
    wb = openpyxl.Workbook()

    # ---- sheet 1: data ----
    ws_data = wb.active
    ws_data.title = "Bragg curves"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center")

    for row in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(row)

    # style the header row
    for cell in ws_data[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # auto-width (capped at 30)
    for col in ws_data.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_data.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    # freeze the header row
    ws_data.freeze_panes = "A2"

    # ---- sheet 2: metadata ----
    ws_meta = wb.create_sheet("Metadata")
    for row in dataframe_to_rows(df_meta, index=False, header=True):
        ws_meta.append(row)

    for cell in ws_meta[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for col in ws_meta.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_meta.column_dimensions[col[0].column_letter].width = min(max_len + 2, 35)

    wb.save(path)
